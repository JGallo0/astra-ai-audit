"""
Co2mply — Serviços auxiliares do módulo Viabilidade:
  1. Extração de premissas de planilha via LLM
  2. Geração de workbook Excel
"""
from __future__ import annotations
import io
import json
from datetime import datetime
from typing import Any

import pandas as pd


# ── Extração via LLM ──────────────────────────────────────────────────────────
# Todos os campos usam nomes canônicos de PremissasViabilidade (sem sufixo _brl)

EXTRACTION_SCHEMA = {
    "feedstock_t_ano":       "Toneladas de feedstock/biomassa seca processada por ano (t/ano)",
    "yield_pirolise":        "Rendimento da pirólise em fração decimal (ex: 28% → 0.28)",
    "fator_carbono":         "Fator de carbono: tCO₂e por tonelada de biochar produzido",
    "preco_credito_usd":     "Preço assumido do crédito de carbono em USD por tCO₂e",
    "fx_rate":               "Taxa de câmbio da moeda do crédito (USD) para a moeda do projeto",
    "preco_biochar":         "Preço de venda do biochar na moeda do projeto por tonelada (0 se não vendido)",
    "capex_total":           "CAPEX total do projeto na moeda do projeto",
    "opex_anual":            "OPEX (custos operacionais totais) anuais na moeda do projeto",
    "wacc":                  "Taxa de desconto ou WACC em fração decimal (ex: 12% → 0.12)",
    "aliquota_efetiva_ir":   "Alíquota efetiva de imposto de renda em fração decimal (ex: 20% → 0.20)",
    "horizonte_anos":        "Horizonte do projeto em anos (geralmente 20)",
    "ano_investimento":      "Ano do investimento inicial (ex: 2026)",
    "escalacao_carbono":     "Escalação anual do preço do carbono em fração (ex: 3% → 0.03)",
    "escalacao_opex":        "Escalação anual do OPEX em fração (ex: 5% → 0.05)",
}


def extract_premissas_from_spreadsheet(
    file_bytes: bytes,
    filename: str,
    openai_client: Any,
    model: str,
) -> dict:
    """
    Lê o arquivo (xlsx/csv), converte para texto e usa LLM para extrair premissas.
    Retorna dict com nomes canônicos. None para campos não encontrados.
    """
    try:
        if filename.lower().endswith(".csv"):
            df_dict = {"Principal": pd.read_csv(io.BytesIO(file_bytes))}
        else:
            df_dict = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    except Exception as e:
        return {"_erro": f"Não foi possível ler o arquivo: {e}"}

    parts = []
    for sheet, df in df_dict.items():
        parts.append(f"\n=== Aba: {sheet} ===")
        parts.append(df.head(60).to_string(max_cols=20))
    text = "\n".join(parts)[:14_000]

    schema_text = "\n".join(f'  "{k}": {v}' for k, v in EXTRACTION_SCHEMA.items())

    prompt = f"""Analise esta planilha de modelagem financeira de projeto de biochar e extraia os parâmetros listados.

PLANILHA:
{text}

PARÂMETROS A EXTRAIR:
{{{schema_text}}}

REGRAS:
- Retorne apenas JSON válido, sem markdown
- Use null para campos não encontrados — nunca invente valores
- Valores percentuais devem ser convertidos para fração decimal (28% → 0.28)
- opex_anual deve ser o total de TODOS os custos operacionais anuais somados
- capex_total deve ser o investimento total inicial (físico + intangível)
- Se encontrar múltiplos cenários, use o cenário base ou mais conservador"""

    resp = openai_client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "Especialista em extração de dados financeiros de planilhas. Retorne apenas JSON válido."},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )

    try:
        extracted = json.loads(resp.choices[0].message.content)
    except Exception:
        extracted = {}

    valid_keys = set(EXTRACTION_SCHEMA.keys())
    return {k: v for k, v in extracted.items() if k in valid_keys and v is not None}


# ── Geração de Excel ──────────────────────────────────────────────────────────

def _fmt_moeda(v, symbol="R$"):
    if v is None: return "—"
    return f"{symbol} {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

def _fmt_pct(v):
    if v is None: return "—"
    return f"{v:.1f}%"

def _fmt_usd(v):
    if v is None: return "—"
    return f"$ {v:,.0f}".replace(",", ".")


def _compute_dre_anual(premissas: dict) -> list[dict]:
    """Reconstrói DRE anual a partir das premissas — para o Excel FCL detalhado."""
    feedstock  = premissas.get("feedstock_t_ano", 0) or 0
    yld        = premissas.get("yield_pirolise", 0.28) or 0.28
    fc         = premissas.get("fator_carbono", 2.5) or 2.5
    preco      = premissas.get("preco_credito_usd", 0) or 0
    fx         = premissas.get("fx_rate", 1.0) or 1.0
    biochar_p  = premissas.get("preco_biochar", 0) or 0
    capex      = premissas.get("capex_total", 0) or 0
    opex       = premissas.get("opex_anual", 0) or 0
    vida       = premissas.get("vida_util_anos", 20) or 20
    aliq       = premissas.get("aliquota_efetiva_ir", 0.20) or 0.0
    horizonte  = premissas.get("horizonte_anos", 20) or 20
    esc_carb   = premissas.get("escalacao_carbono", 0) or 0
    esc_fx     = premissas.get("escalacao_fx", 0) or 0
    esc_opex   = premissas.get("escalacao_opex", 0) or 0
    ano_inv    = premissas.get("ano_investimento", 2026) or 2026

    biochar  = feedstock * yld
    creditos = biochar * fc
    da       = capex / max(vida, 1)
    rows     = []

    for i in range(1, horizonte + 1):
        ec  = (1 + esc_carb) ** (i - 1)
        ef  = (1 + esc_fx)   ** (i - 1)
        eo  = (1 + esc_opex) ** (i - 1)
        rec     = creditos * preco * ec * fx * ef + biochar * biochar_p
        opex_i  = opex * eo
        ebitda  = rec - opex_i
        ebit    = ebitda - da
        trib    = max(ebit, 0) * aliq
        fcl_i   = ebit - trib + da
        rows.append({
            "ano":     ano_inv + i,
            "receita": round(rec, 0),
            "opex":    round(opex_i, 0),
            "ebitda":  round(ebitda, 0),
            "da":      round(da, 0),
            "ebit":    round(ebit, 0),
            "trib":    round(trib, 0),
            "fcl":     round(fcl_i, 0),
        })
    return rows


def generate_viabilidade_excel(premissas: dict, resultado: dict, project_name: str) -> bytes:
    """Gera workbook Excel com 4 abas: Premissas, Indicadores, DRE+FCL, Sensibilidade."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl não instalado.")

    NAVY  = "1A3160"
    GREEN = "16A34A"
    AMBER = "B45309"
    RED   = "DC2626"
    LGRAY = "F3F4F6"
    WHITE = "FFFFFF"
    DGRAY = "374151"

    def hdr_fill(color=NAVY):  return PatternFill("solid", fgColor=color)
    def hdr_font(bold=True, color=WHITE): return Font(bold=bold, color=color, name="Calibri", size=11)
    def data_font(bold=False, color="000000"): return Font(bold=bold, color=color, name="Calibri", size=10)

    def bdr():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_row(ws, row, label, value, fmt="text"):
        cl = ws.cell(row=row, column=1, value=label)
        cl.font = data_font(); cl.fill = PatternFill("solid", fgColor=LGRAY)
        cl.border = bdr(); cl.alignment = Alignment(horizontal="left")
        cv = ws.cell(row=row, column=2, value=value)
        cv.font = data_font(); cv.border = bdr()
        cv.alignment = Alignment(horizontal="right")
        if fmt == "num" and isinstance(value, (int, float)):
            cv.number_format = '#,##0'
        elif fmt == "pct" and isinstance(value, (int, float)):
            cv.number_format = '0.00%'
        elif fmt == "usd" and isinstance(value, (int, float)):
            cv.number_format = '"$"#,##0.00'

    def section(ws, row, title, ncols=2):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
        c.fill = hdr_fill(DGRAY)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[row].height = 20

    wb   = Workbook()
    date_str = datetime.now().strftime("%d/%m/%Y")

    # moeda do projeto
    moeda   = premissas.get("moeda_projeto", "BRL")
    sym     = {"BRL":"R$","USD":"$","EUR":"€","GBP":"£"}.get(moeda, moeda)
    wacc_v  = premissas.get("wacc", 0.12) or 0.12

    # ── Aba 1: Premissas ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Premissas"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 22

    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = f"Co2mply — Premissas de Viabilidade | {project_name} | {date_str}"
    c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
    c.fill = hdr_fill(NAVY); c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    r = 2
    section(ws1, r, "PRODUÇÃO"); r += 1
    set_row(ws1, r, "Feedstock (t/ano, base seca)",       premissas.get("feedstock_t_ano"),   "num"); r += 1
    set_row(ws1, r, "Rendimento de pirólise",              premissas.get("yield_pirolise"),    "pct"); r += 1
    set_row(ws1, r, "Fator de carbono (tCO₂/t biochar)",  premissas.get("fator_carbono")); r += 1
    set_row(ws1, r, "Biochar produzido (t/ano)",           resultado.get("biochar_t_ano"),     "num"); r += 1
    set_row(ws1, r, "Créditos gerados (tCO₂e/ano)",       resultado.get("creditos_tco2_ano"), "num"); r += 1

    section(ws1, r, "RECEITAS"); r += 1
    set_row(ws1, r, "Preço crédito (USD/tCO₂e)",          premissas.get("preco_credito_usd"), "usd"); r += 1
    set_row(ws1, r, f"Câmbio USD → {moeda}",               premissas.get("fx_rate")); r += 1
    set_row(ws1, r, f"Preço biochar ({sym}/t)",            premissas.get("preco_biochar"),     "num"); r += 1
    set_row(ws1, r, f"Receita bruta ano 1 ({sym})",        resultado.get("receita_bruta_yr1"), "num"); r += 1

    section(ws1, r, "CUSTOS"); r += 1
    set_row(ws1, r, f"CAPEX total ({sym})",                premissas.get("capex_total"),       "num"); r += 1
    set_row(ws1, r, f"OPEX anual ({sym})",                 premissas.get("opex_anual"),        "num"); r += 1
    set_row(ws1, r, f"Depreciação anual ({sym})",          resultado.get("da_anual"),          "num"); r += 1
    set_row(ws1, r, f"EBITDA ano 1 ({sym})",               resultado.get("ebitda_yr1"),        "num"); r += 1

    section(ws1, r, "FINANCEIRO"); r += 1
    set_row(ws1, r, "WACC / Taxa de desconto",             premissas.get("wacc"),              "pct"); r += 1
    set_row(ws1, r, "Alíquota efetiva de IR",              premissas.get("aliquota_efetiva_ir"),"pct"); r += 1
    set_row(ws1, r, "Horizonte (anos)",                    premissas.get("horizonte_anos")); r += 1
    set_row(ws1, r, "Ano de investimento",                 premissas.get("ano_investimento")); r += 1
    set_row(ws1, r, "Moeda do projeto",                    moeda); r += 1

    # ── Aba 2: Indicadores ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Indicadores")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 22

    ws2.merge_cells("A1:B1")
    c = ws2["A1"]
    c.value = "INDICADORES FINANCEIROS"
    c.font = hdr_font(); c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    irr    = resultado.get("irr")
    irr_sc = resultado.get("irr_sem_carbono")
    npv    = resultado.get("npv")   # campo canônico

    indicators = [
        ("TIR (IRR)",              f"{irr:.1f}%" if irr is not None else "—"),
        (f"VPL ({sym})",           f"{sym} {npv:,.0f}" if npv is not None else "—"),
        ("Payback",                str(resultado.get("payback_year")) if resultado.get("payback_year") else "Não atingido"),
        (f"EBITDA Ano 1 ({sym})",  f"{sym} {resultado.get('ebitda_yr1'):,.0f}" if resultado.get('ebitda_yr1') is not None else "—"),
        ("Margem EBITDA Ano 1",    _fmt_pct(resultado.get("margem_ebitda_pct"))),
        ("—", "—"),
        ("ADICIONALIDADE FINANCEIRA", ""),
        ("TIR sem receita de carbono", f"{irr_sc:.1f}%" if irr_sc is not None else "Inviável (IRR < 0)"),
        ("Adicionalidade confirmada?",  "✓ SIM" if resultado.get("adicionalidade_financeira") else "✗ NÃO"),
        ("—", "—"),
        ("Break-even (USD/tCO₂)",  _fmt_usd(resultado.get("preco_breakeven_usd"))),
    ]

    for i, (lbl, val) in enumerate(indicators, start=2):
        cl = ws2.cell(row=i, column=1, value=lbl)
        cv = ws2.cell(row=i, column=2, value=val)
        cl.font = Font(bold=lbl in ("ADICIONALIDADE FINANCEIRA",), name="Calibri", size=10)
        cl.fill = PatternFill("solid", fgColor=LGRAY)
        cv.font = Font(bold=True, name="Calibri", size=10)
        for cell in (cl, cv):
            cell.border = bdr()
            cell.alignment = Alignment(horizontal="right" if cell.column == 2 else "left")

    # ── Aba 3: DRE + Fluxo de Caixa ──────────────────────────────────────────
    ws3 = wb.create_sheet("DRE + Fluxo de Caixa")
    dre = _compute_dre_anual(premissas)
    n_anos = len(dre)

    # Ajusta larguras
    ws3.column_dimensions["A"].width = 26
    for col in range(2, n_anos + 3):
        ws3.column_dimensions[get_column_letter(col)].width = 13

    # Título
    ws3.merge_cells(f"A1:{get_column_letter(n_anos + 1)}1")
    c = ws3["A1"]
    c.value = f"DRE E FLUXO DE CAIXA LIVRE — {n_anos} ANOS ({moeda})"
    c.font = hdr_font(); c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 24

    # Header anos
    header = [""] + [str(d["ano"]) for d in dre]
    for col, val in enumerate(header, start=1):
        c = ws3.cell(row=2, column=col, value=val)
        c.font = hdr_font(color=WHITE); c.fill = hdr_fill(DGRAY)
        c.alignment = Alignment(horizontal="center")

    # Linha de CAPEX (ano 0 = -capex)
    capex_v = premissas.get("capex_total", 0) or 0
    capex_row = ["(-) CAPEX"] + [None] * n_anos
    capex_row[1] = -capex_v   # ano 0 (col 2)

    rows_dre = [
        ("Receita Bruta",       [d["receita"] for d in dre],  GREEN),
        ("(-) OPEX",            [-d["opex"]   for d in dre],  RED),
        ("= EBITDA",            [d["ebitda"]  for d in dre],  NAVY),
        ("(-) Depreciação",     [-d["da"]     for d in dre],  DGRAY),
        ("= EBIT",              [d["ebit"]    for d in dre],  NAVY),
        ("(-) Impostos",        [-d["trib"]   for d in dre],  DGRAY),
        ("= FCL Anual",         [d["fcl"]     for d in dre],  NAVY),
    ]

    # CAPEX row first
    for col, val in enumerate(capex_row, start=1):
        c = ws3.cell(row=3, column=col, value=val)
        c.font = Font(bold=True, color=RED if isinstance(val, (int, float)) and val < 0 else "000000", name="Calibri", size=10)
        c.border = bdr(); c.alignment = Alignment(horizontal="right" if col > 1 else "left")
        if isinstance(val, (int, float)): c.number_format = '#,##0'

    # DRE rows
    for row_i, (label, values, color) in enumerate(rows_dre, start=4):
        is_total = label.startswith("=")
        c = ws3.cell(row=row_i, column=1, value=label)
        c.font = Font(bold=is_total, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor="EEF2FA" if is_total else LGRAY)
        c.border = bdr()
        for col_i, val in enumerate(values, start=2):
            cell = ws3.cell(row=row_i, column=col_i, value=val)
            col_r = color if val >= 0 else RED
            cell.font = Font(bold=is_total, color=col_r, name="Calibri", size=10)
            cell.border = bdr()
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'

    # FCL Acumulado
    acum  = resultado.get("fcl_acumulado", [])
    anos_labels = resultado.get("anos", [])
    acum_row_data = ["FCL Acumulado"] + [float(v) for v in acum]
    row_acum = 4 + len(rows_dre)
    for col, val in enumerate(acum_row_data, start=1):
        c = ws3.cell(row=row_acum, column=col, value=val)
        c.font = Font(bold=True, color=GREEN if isinstance(val, (int,float)) and val >= 0 else RED, name="Calibri", size=10)
        c.border = bdr(); c.alignment = Alignment(horizontal="right" if col > 1 else "left")
        if isinstance(val, float): c.number_format = '#,##0'

    # ── Aba 4: Sensibilidade ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Sensibilidade")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 18

    ws4.merge_cells("A1:C1")
    c = ws4["A1"]
    c.value = "SENSIBILIDADE — Preço do Crédito × TIR / VPL"
    c.font = hdr_font(); c.fill = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Preço Crédito (USD)", "TIR (%)", f"VPL ({moeda})"], start=1):
        c = ws4.cell(row=2, column=col, value=h)
        c.font = hdr_font(color=WHITE); c.fill = hdr_fill(DGRAY)
        c.alignment = Alignment(horizontal="center")

    for row_i, s in enumerate(resultado.get("sensibilidade", []), start=3):
        irr_s = s.get("irr")
        npv_s = s.get("npv")   # campo canônico
        for col, val in enumerate([s["preco_usd"], irr_s, npv_s], start=1):
            c = ws4.cell(row=row_i, column=col, value=val)
            c.border = bdr(); c.alignment = Alignment(horizontal="right")
            if col == 1:
                c.number_format = '"$"#,##0'
                c.font = data_font()
            elif col == 2 and irr_s is not None:
                ok = irr_s >= (wacc_v * 100)
                c.font = Font(color=GREEN if ok else RED, name="Calibri", size=10)
                c.number_format = '0.00'
            elif col == 3 and npv_s is not None:
                c.font = Font(color=GREEN if npv_s >= 0 else RED, name="Calibri", size=10)
                c.number_format = '#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Memo Financeiro PDF ───────────────────────────────────────────────────────

def generate_financial_memo_pdf(premissas: dict, resultado: dict, project_name: str) -> bytes:
    """
    Memo financeiro executivo — 2 páginas, para apresentação a investidores.
    Usa ReportLab com o mesmo design system dos outros PDFs do Co2mply.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether,
    )
    import os

    NAVY   = "#1A3160"; NAVY_L = "#EEF2FA"
    GREEN  = "#16A34A"; GREEN_B = "#F0FDF4"
    AMBER  = "#B45309"; AMBER_B = "#FFFBEB"
    RED    = "#DC2626"; RED_B   = "#FEF2F2"
    GRAY   = "#6B7280"; BORDER  = "#E5E7EB"
    TEXT   = "#111827"; TEXT2   = "#6B7280"
    WHITE  = "#FFFFFF"

    def _c(h): return colors.HexColor(h)
    def _p(text, **kw):
        base = dict(fontName="Helvetica", fontSize=9, leading=12, textColor=_c(TEXT))
        base.update(kw)
        return Paragraph(text, ParagraphStyle("s", **base))
    def _s(n=0.3): return Spacer(1, n * cm)

    PAGE_W, PAGE_H = A4
    ML = 1.8 * cm; MR = 1.8 * cm; MT = 2.0 * cm; MB = 1.8 * cm
    TW = PAGE_W - ML - MR
    date_str = datetime.now().strftime("%d/%m/%Y")
    moeda = premissas.get("moeda_projeto", "BRL")
    sym   = {"BRL":"R$","USD":"$","EUR":"€","GBP":"£"}.get(moeda, moeda)

    # ── Assets ──────────────────────────────────────────────────────────────────
    _ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    LOGO_PATH  = os.path.join(_ROOT, "assets", "logo_transparent.png")
    LOGO_RATIO = 743 / 320

    def _page_cb(canvas, doc):
        canvas.saveState()
        HDR_H = 1.5 * cm
        canvas.setFillColor(_c("#F1F5F9"))
        canvas.rect(0, PAGE_H - HDR_H, PAGE_W, HDR_H, fill=1, stroke=0)
        canvas.setStrokeColor(_c(NAVY)); canvas.setLineWidth(1.2)
        canvas.line(0, PAGE_H - HDR_H, PAGE_W, PAGE_H - HDR_H)
        PAD = 0.15 * cm
        LH = HDR_H - 2 * PAD; LW = LH * LOGO_RATIO
        if os.path.exists(LOGO_PATH):
            canvas.drawImage(LOGO_PATH, ML, PAGE_H - HDR_H + PAD,
                             width=LW, height=LH, preserveAspectRatio=True, mask="auto")
        canvas.setFillColor(_c(NAVY)); canvas.setFont("Helvetica-Bold", 8)
        canvas.drawRightString(PAGE_W - MR, PAGE_H - HDR_H/2 + 0.1*cm, "Co2mply")
        canvas.setFont("Helvetica", 7); canvas.setFillColor(_c(GRAY))
        canvas.drawRightString(PAGE_W - MR, PAGE_H - HDR_H/2 - 0.22*cm, "by Astra Carbon")
        # Footer
        canvas.setStrokeColor(_c(BORDER)); canvas.setLineWidth(0.5)
        canvas.line(ML, 1.4*cm, PAGE_W - MR, 1.4*cm)
        canvas.setFillColor(_c(TEXT2)); canvas.setFont("Helvetica", 7)
        canvas.drawString(ML, 0.8*cm, f"Memo Financeiro — {project_name} | CONFIDENCIAL")
        canvas.drawRightString(PAGE_W - MR, 0.8*cm, f"Página {doc.page} | {date_str}")
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=ML, rightMargin=MR,
        topMargin=MT + 1.6*cm, bottomMargin=MB + 1.0*cm,
        title=f"Co2mply | Memo Financeiro — {project_name}",
    )
    story = []

    # ── Título ───────────────────────────────────────────────────────────────────
    story.append(_p("MEMO FINANCEIRO — ANÁLISE DE VIABILIDADE",
                    fontName="Helvetica-Bold", fontSize=8, textColor=_c(TEXT2),
                    letterSpacing=1.2))
    story.append(_p(project_name,
                    fontName="Helvetica-Bold", fontSize=22, leading=28, textColor=_c(NAVY)))
    story.append(_p(f"Metodologia: {premissas.get('metodologia', 'Biochar CDR')}  ·  "
                    f"Moeda: {moeda}  ·  Data: {date_str}",
                    fontSize=9, textColor=_c(TEXT2)))
    story.append(_s(0.3))
    story.append(HRFlowable(width="100%", thickness=2.5, color=_c(NAVY)))
    story.append(_s(0.5))

    # ── KPI block ─────────────────────────────────────────────────────────────────
    irr    = resultado.get("irr")
    irr_eq = resultado.get("irr_equity")
    npv    = resultado.get("npv")
    pb     = resultado.get("payback_year")
    wacc   = premissas.get("wacc", 0.12) * 100
    fpc    = premissas.get("financiamento_pct", 0) or 0

    def _kpi_color(val, wacc_pct):
        if val is None: return GRAY
        return GREEN if val >= wacc_pct else RED

    kpi_cols = [
        (f"{irr:.1f}%" if irr else "—", "TIR do Projeto",
         _kpi_color(irr, wacc), f"WACC ref.: {wacc:.0f}%"),
        (f"{irr_eq:.1f}%" if irr_eq else ("Sem alavancagem" if fpc == 0 else "—"),
         "TIR do Equity",
         _kpi_color(irr_eq, wacc) if irr_eq else GRAY,
         f"{fpc*100:.0f}% dívida" if fpc > 0 else "Projeto 100% equity"),
        (f"{sym} {npv/1e6:.1f}M" if npv else "—", "VPL",
         GREEN if (npv or 0) >= 0 else RED, f"Taxa: {wacc:.0f}%"),
        (str(pb) if pb else "Não atingido", "Payback",
         GREEN if pb else AMBER, "Ano de retorno"),
    ]

    c_w = TW / len(kpi_cols)
    kpi_tbl = Table(
        [[_p(v, fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=_c(col),
              alignment=TA_CENTER) for v, _, col, _ in kpi_cols],
         [_p(lbl, fontSize=8, textColor=_c(TEXT2), alignment=TA_CENTER)
          for _, lbl, _, _ in kpi_cols],
         [_p(sub, fontSize=7.5, textColor=_c(GRAY), alignment=TA_CENTER)
          for _, _, _, sub in kpi_cols]],
        colWidths=[c_w] * len(kpi_cols),
        rowHeights=[1.2*cm, 0.55*cm, 0.45*cm],
    )
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), _c(NAVY_L)),
        ("ALIGN",  (0,0),(-1,-1), "CENTER"), ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,0),10), ("BOTTOMPADDING",(0,0),(-1,0),4),
        ("TOPPADDING",(0,1),(-1,-1),3), ("BOTTOMPADDING",(0,1),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),4), ("RIGHTPADDING",(0,0),(-1,-1),4),
        ("LINEAFTER",(0,0),(2,-1),0.5,_c(BORDER)),
        ("BOX",(0,0),(-1,-1),0.5,_c(BORDER)),
    ]))
    story.append(kpi_tbl)
    story.append(_s(0.5))

    # ── Premissas principais ──────────────────────────────────────────────────────
    story.append(_p("Premissas Principais",
                    fontName="Helvetica-Bold", fontSize=11, textColor=_c(NAVY)))
    story.append(_s(0.2))

    def prow(lbl, val):
        return [_p(lbl, fontSize=9, textColor=_c(TEXT2)),
                _p(str(val), fontName="Helvetica-Bold", fontSize=9, alignment=TA_RIGHT)]

    prem_data = [
        prow("Feedstock (t/ano)", f"{premissas.get('feedstock_t_ano',0):,.0f} t"),
        prow("Yield pirólise", f"{premissas.get('yield_pirolise',0)*100:.0f}%"),
        prow("Fator carbono (tCO₂/t biochar)", f"{premissas.get('fator_carbono',0):.2f}"),
        prow("Preço do crédito", f"$ {premissas.get('preco_credito_usd',0):.0f} / tCO₂e"),
        prow(f"Câmbio USD→{moeda}", f"{premissas.get('fx_rate',1):.2f}"),
        prow(f"CAPEX total ({sym})", f"{sym} {premissas.get('capex_total',0):,.0f}"),
        prow(f"OPEX anual ({sym})", f"{sym} {premissas.get('opex_anual',0):,.0f}"),
        prow("WACC", f"{premissas.get('wacc',0.12)*100:.1f}%"),
        prow("Alíquota efetiva IR", f"{premissas.get('aliquota_efetiva_ir',0.20)*100:.0f}%"),
        prow("Buffer pool", f"{premissas.get('buffer_pool_pct',0)*100:.1f}%"),
        prow("Delay emissão", f"{premissas.get('issuance_delay_months',0)} meses"),
    ]

    if fpc > 0:
        prem_data += [
            prow("Financiamento (% CAPEX)", f"{fpc*100:.0f}%"),
            prow("Taxa de juros", f"{premissas.get('taxa_juros',0.12)*100:.1f}%"),
            prow("Prazo financiamento", f"{premissas.get('prazo_financiamento',10)} anos"),
        ]

    # Split into 2 columns
    half = len(prem_data) // 2 + len(prem_data) % 2
    col1 = prem_data[:half]; col2 = prem_data[half:]
    while len(col2) < len(col1): col2.append([_p(""), _p("")])

    prem_merged = []
    CW = TW / 2 - 0.3*cm
    for r1, r2 in zip(col1, col2):
        prem_merged.append(r1 + [_p("")] + r2)

    prem_tbl = Table(prem_merged, colWidths=[CW*0.62, CW*0.38, 0.6*cm, CW*0.62, CW*0.38])
    prem_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(1,-1),_c("#F9FAFB")),
        ("BACKGROUND",(3,0),(4,-1),_c("#F9FAFB")),
        ("LINEBELOW",(0,0),(1,-1),0.3,_c(BORDER)),
        ("LINEBELOW",(3,0),(4,-1),0.3,_c(BORDER)),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6),
    ]))
    story.append(prem_tbl)
    story.append(_s(0.5))

    # ── DRE Resumida ──────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=_c(BORDER)))
    story.append(_s(0.3))
    story.append(_p("Demonstração de Resultados — Anos Selecionados",
                    fontName="Helvetica-Bold", fontSize=11, textColor=_c(NAVY)))
    story.append(_s(0.2))

    from backend.viabilidade_service import _compute_dre_anual
    dre = _compute_dre_anual(premissas)
    sel_anos = [0, 1, 2, 4, 9, 19]  # índices 0-based → anos 1,2,3,5,10,20
    sel = [dre[i] for i in sel_anos if i < len(dre)]

    ano_hdrs = [""] + [str(d["ano"]) for d in sel]
    rows_dre = [
        ("Receita bruta",  [d["receita"] for d in sel],  GREEN),
        ("(−) OPEX",       [-d["opex"]   for d in sel],  RED),
        ("= EBITDA",       [d["ebitda"]  for d in sel],  NAVY),
        ("(−) DA",         [-d["da"]     for d in sel],  GRAY),
        ("= EBIT",         [d["ebit"]    for d in sel],  NAVY),
        ("(−) Impostos",   [-d["trib"]   for d in sel],  GRAY),
        ("= FCL",          [d["fcl"]     for d in sel],  NAVY),
    ]

    def _fval(v):
        if v is None: return "—"
        m = abs(v) / 1e6
        return f"{'-' if v < 0 else ''}{sym}{m:.2f}M"

    dre_data = [ano_hdrs]
    for label, vals, col in rows_dre:
        is_tot = label.startswith("=")
        row = [_p(label, fontName="Helvetica-Bold" if is_tot else "Helvetica",
                  fontSize=8, textColor=_c(NAVY if is_tot else TEXT2))]
        for v in vals:
            c_str = GREEN if v >= 0 else RED
            row.append(_p(_fval(v), fontName="Helvetica-Bold" if is_tot else "Helvetica",
                          fontSize=8, textColor=_c(c_str if is_tot else TEXT),
                          alignment=TA_RIGHT))
        dre_data.append(row)

    ncols = len(sel) + 1
    dre_cw = [3.0*cm] + [(TW - 3.0*cm) / len(sel)] * len(sel)
    dre_tbl = Table(dre_data, colWidths=dre_cw)
    dre_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),_c("#374151")),
        ("FONTCOLOR",(0,0),(-1,0),_c(WHITE)),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5), ("RIGHTPADDING",(0,0),(-1,-1),5),
        ("LINEBELOW",(0,0),(-1,-1),0.3,_c(BORDER)),
        ("BACKGROUND",(0,3),(0,3),_c(NAVY_L)),  # EBITDA
        ("BACKGROUND",(0,5),(0,5),_c(NAVY_L)),  # EBIT
        ("BACKGROUND",(0,7),(0,7),_c(NAVY_L)),  # FCL
    ]))
    story.append(dre_tbl)
    story.append(_s(0.5))

    # ── Sensibilidade compacta ────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=_c(BORDER)))
    story.append(_s(0.3))
    story.append(_p("Análise de Sensibilidade — Preço do Crédito",
                    fontName="Helvetica-Bold", fontSize=11, textColor=_c(NAVY)))
    story.append(_s(0.2))

    sensi = resultado.get("sensibilidade", [])
    # Seleciona ~10 pontos espaçados
    step = max(1, len(sensi) // 10)
    sel_s = sensi[::step][:10]
    if sel_s:
        sens_hdr = [_p("USD/tCO₂", fontSize=8, fontName="Helvetica-Bold", alignment=TA_CENTER)] + \
                   [_p(f"${s['preco_usd']}", fontSize=8, alignment=TA_CENTER) for s in sel_s]
        wacc_pct = premissas.get("wacc", 0.12) * 100
        irr_row  = [_p("TIR (%)", fontSize=8, fontName="Helvetica-Bold")] + \
                   [_p(f"{s['irr']:.1f}%" if s['irr'] else "—", fontSize=8,
                       fontName="Helvetica-Bold",
                       textColor=_c(GREEN if (s['irr'] or 0) >= wacc_pct else RED),
                       alignment=TA_CENTER) for s in sel_s]

        cw_s = [2.5*cm] + [(TW - 2.5*cm) / len(sel_s)] * len(sel_s)
        sens_tbl = Table([sens_hdr, irr_row], colWidths=cw_s)
        sens_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),_c("#374151")),
            ("FONTCOLOR",(0,0),(-1,0),_c(WHITE)),
            ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),4), ("RIGHTPADDING",(0,0),(-1,-1),4),
            ("BACKGROUND",(0,1),(-1,1),_c(NAVY_L)),
            ("BOX",(0,0),(-1,-1),0.5,_c(BORDER)),
        ]))
        story.append(sens_tbl)
        story.append(_s(0.3))
        story.append(_p(f"Verde = TIR ≥ WACC ({wacc_pct:.0f}%). Break-even: "
                        f"${resultado.get('preco_breakeven_usd','—')}/tCO₂.",
                        fontSize=8, textColor=_c(GRAY)))

    # ── Disclaimer ───────────────────────────────────────────────────────────────
    story.append(_s(0.5))
    story.append(HRFlowable(width="100%", thickness=0.4, color=_c(BORDER)))
    story.append(_s(0.2))
    story.append(_p(
        f"Este memo foi gerado pelo motor determinístico Co2mply para o projeto <b>{project_name}</b> "
        f"em {date_str}. Os indicadores refletem as premissas informadas pelo proponente e não "
        f"constituem rating, recomendação de investimento ou garantia de resultado. "
        f"Co2mply by Astra Carbon · v1.0 · CONFIDENCIAL",
        fontSize=7, textColor=_c(GRAY), leading=10,
        fontName="Helvetica-Oblique",
    ))

    doc.build(story, onFirstPage=_page_cb, onLaterPages=_page_cb)
    return buf.getvalue()
